from django.shortcuts import render
from django.http import HttpResponse
from goods.models import Products
from main.forms import ExportForm
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill, NamedStyle
from openpyxl.styles.colors import COLOR_INDEX
from typing import List, Dict, Union
from datetime import datetime
from urllib.parse import quote


# Create your views here.
def index(request):

    context = {
        'title': 'Главная',
        'content': 'Магазин женской одежды - Пассаж',
    }
    return render(request, 'main/index.html', context)


def about(request):
    context = {
        'title': 'Главная - О нас',
        'content': 'О нас',
    }
    return render(request, 'main/about.html', context)


def panel(request):
    if request.method == 'POST':
        form = ExportForm(request.POST)
        if form.is_valid():
            category_name = form.cleaned_data['category']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']

            products = Products.objects.filter(category__name=category_name)

            if start_date:
                products = products.filter(created_at__gte=start_date)
            if end_date:
                products = products.filter(created_at__lte=end_date)

            wb: Workbook = Workbook()
            ws = wb.active
            ws.title = f'Товары'

            headers: List[Dict[str, Union[str, int]]] = [
                {'name': 'code', 'title': 'Код товара', 'width': 20, 'stroka2': 'GTIN', 'stroka3': 'value', 'stroka4': ''},
                {'name': 'tnvd', 'title': 'Код ТНВЭД', 'width': 20, 'stroka2': 'Tnved', 'stroka3': 'value', 'stroka4': ''},
                {'name': 'name', 'title': 'Полное наименование товара', 'width': 40, 'stroka2': '2478', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'znak', 'title': 'Товарный знак', 'width': 30, 'stroka2': '2504', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'article', 'title': 'Артикул производителя', 'width': 30, 'stroka2': '13914', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'category', 'title': 'Вид товара', 'width': 20, 'stroka2': '12', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'color', 'title': 'Цвет', 'width': 20, 'stroka2': '36', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'pol', 'title': 'Целевой пол', 'width': 20, 'stroka2': '14013', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'size', 'title': 'Размер одежды изделия', 'width': 20, 'stroka2': '35', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'consist', 'title': 'Состав', 'width': 20, 'stroka2': '2483', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'codetnvd', 'title': 'Код ТНВЭД', 'width': 20, 'stroka2': '13933', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'reglament', 'title': 'Номер технического регламента', 'width': 60, 'stroka2': '13836', 'stroka3': 'value', 'stroka4': 'Текстовое значение'},
                {'name': 'status', 'title': 'Статус карточки товара в Каталоге', 'width': 40, 'stroka2': 'status', 'stroka3': 'value', 'stroka4': 'Текстовое поле(Черновик или На модерации)'},
                {'name': 'result', 'title': 'Результат обработки данных в Каталоге', 'width': 40, 'stroka2': 'result', 'stroka3': 'value', 'stroka4': 'Заполняется автоматически при загрузке в систему'}
            ]

            table_style = NamedStyle(name='table_style')
            table_style.font = Font(name='Arial')

            for position, header in enumerate(headers):
                ws.cell(1, position + 1, header['title']).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(1, position + 1, header['title']).font = Font(name='Arial', bold=True)
                ws.cell(1, position + 1, header['title']).fill = PatternFill(fgColor=COLOR_INDEX[5], fill_type="solid")
                ws.cell(1, position + 1, header['title']).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                ws.cell(2, position + 1, header['stroka2']).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(2, position + 1, header['stroka2']).font = Font(name='Arial')
                ws.cell(2, position + 1, header['stroka2']).fill = PatternFill(fgColor='E3E1E4', fill_type="solid")
                ws.cell(2, position + 1, header['stroka2']).border = Border(left=Side(style='thin'),
                                                                          right=Side(style='thin'),
                                                                          top=Side(style='thin'),
                                                                          bottom=Side(style='thin'))

                ws.cell(3, position + 1, header['stroka3']).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(3, position + 1, header['stroka3']).font = Font(name='Arial')
                ws.cell(3, position + 1, header['stroka3']).fill = PatternFill(fgColor='E3E1E4', fill_type="solid")
                ws.cell(3, position + 1, header['stroka3']).border = Border(left=Side(style='thin'),
                                                                          right=Side(style='thin'),
                                                                          top=Side(style='thin'),
                                                                          bottom=Side(style='thin'))

                ws.cell(4, position + 1, header['stroka4']).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(4, position + 1, header['stroka4']).font = Font(name='Arial')
                ws.cell(4, position + 1, header['stroka4']).fill = PatternFill(fgColor='E3E1E4', fill_type="solid")
                ws.cell(4, position + 1, header['stroka4']).border = Border(left=Side(style='thin'),
                                                                          right=Side(style='thin'),
                                                                          top=Side(style='thin'),
                                                                          bottom=Side(style='thin'))

                ws.column_dimensions[ws.cell(1, position + 1).column_letter].width = header['width']

                start_position: int = 5
                for product in products:
                    for sizeproduct in product.sizeproductrelation_set.all():
                        if sizeproduct.count > 0:
                            for position, header in enumerate(headers):
                                if header['name'] == 'tnvd':
                                    ws.cell(start_position, position + 1).value = sizeproduct.product.category.tnved
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'name':
                                    ws.cell(start_position,
                                            position + 1).value = f'{sizeproduct.product.name}, р.{sizeproduct.size.name}, арт.{sizeproduct.product.article}'
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'znak':
                                    ws.cell(start_position, position + 1).value = 'Без товарного знака'
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'article':
                                    ws.cell(start_position, position + 1).value = sizeproduct.product.article
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'category':
                                    ws.cell(start_position, position + 1).value = sizeproduct.product.category.name
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'color':
                                    colors_names = [color.name for color in sizeproduct.product.color.all()]
                                    colors_string = ", ".join(colors_names)
                                    ws.cell(start_position, position + 1).value = colors_string
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'pol':
                                    ws.cell(start_position, position + 1).value = 'ЖЕНСКИЙ'
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'size':
                                    ws.cell(start_position, position + 1).value = sizeproduct.size.name
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'consist':
                                    consists_names = [consist.name for consist in sizeproduct.product.consist.all()]
                                    consists_string = ", ".join(consists_names)
                                    ws.cell(start_position, position + 1).value = consists_string
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'codetnvd':
                                    ws.cell(start_position, position + 1).value = f'{sizeproduct.product.category.tnved}000000'
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'reglament':
                                    ws.cell(start_position, position + 1).value = 'TP TC 017/2011 "О безопасности продукции легкой промышленности"'
                                    ws.cell(start_position, position + 1).style = table_style
                                elif header['name'] == 'status':
                                    ws.cell(start_position, position + 1).value = 'Черновик'
                                    ws.cell(start_position, position + 1).style = table_style
                            start_position += 1

            filename = f'products_{category_name}_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
            filepath = os.path.join('media', filename)
            wb.save(filepath)

            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="{}"'.format(quote(filename))

            return response

    else:
        form = ExportForm()

    context = {
        'title': 'Главная - Панель администратора',
        'content': 'Панель администратора',
        'form': form
    }
    return render(request, 'main/panel.html', context)


